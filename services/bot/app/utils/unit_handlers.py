from typing import Dict, Any, Tuple, List
import pandas as pd
import numpy as np
import matplotlib
import matplotlib.pyplot as plt
import io
import zipfile
import logging
from datetime import datetime
from dataclasses import dataclass
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter

matplotlib.use('Agg')
pd.set_option('future.no_silent_downcasting', True)
logging.basicConfig(level=logging.INFO)

def process_dataframe(df: pd.DataFrame) -> pd.DataFrame|None:
    """
    Основная функция обработки DataFrame для расчета юнит-экономики
    
    Args:
        df: Входной DataFrame с бизнес-метриками
        
    Returns:
        pd.DataFrame: Обработанный DataFrame с расчетными метриками
    """
    try:
        df["Unit"] = "User"
        df["C1"] = df["customers"] / df["users"]
        df["ARPC"] = df["AVP"] * df["APC"]
        df["ARPU"] = df["ARPC"] * df["C1"]
        df["CPA"] = df["TMS"] / df["users"]
        df["CAC"] = df["TMS"] / df["customers"]
        df["CLTV"] = (df["AVP"] - df["COGS"]) * df["APC"] - df["COGS1s"]
        df["LTV"] = df["CLTV"] * df["C1"]
        df["ROI"] = (df["LTV"] - df["CPA"]) / df["CPA"] * 100
        df["UCM"] = df["LTV"] - df["CPA"]
        df["CCM"] = df["CLTV"] - df["CAC"]

        df["Profitable"] = df["UCM"] > 0

        df["Revenue"] = df["ARPU"] * df["users"]
        df["Gross_profit"] = df["CLTV"] * df["customers"]
        df["Margin"] = df["Gross_profit"] - df["TMS"]

        def calculate_required_bep(row: pd.Series):
            ucm = row.get("UCM", 0)
            if ucm > 0:
                return row.get("FC", 0) / ucm
            return None

        df["Required_units_to_BEP"] = df.apply(calculate_required_bep, axis=1)
        
        df["BEP"] = df["Required_units_to_BEP"] * df["UCM"]
        df["Profit"] = df["Margin"] - df["FC"]

        float_cols = df.select_dtypes(include='float').columns
        df[float_cols] = df[float_cols].round(4)

        df.replace([np.inf, -np.inf], np.nan, inplace=True)
        df = df.infer_objects(copy=False)
        df = df.where(pd.notnull(df), None)

        df = df.where(pd.notnull(df), None)[["name", "users", "customers", "AVP", "APC", "TMS", "COGS", "COGS1s", "FC", "C1", "ARPC", "ARPU", "CPA", "CAC", "CLTV", "LTV", "ROI", "UCM", "CCM", "Revenue", "Gross_profit", "Margin", "Required_units_to_BEP", "BEP", "Profit"]]
        df.columns = ["name", "users", "customers", "AVP", "APC", "TMS", "COGS", "COGS1s", "FC", "C1", "ARPC", "ARPU", "CPA", "CAC", "CLTV", "LTV", "ROI", "UCM", "CCM", "Revenue", "Gross_profit", "Margin", "Required_units_to_BEP", "BEP", "Profit"]

        return df

    except Exception as e:
        logging.error(f"Ошибка в process_dataframe: {e}")
        return None


def unit_calculate_economics(data: Dict[str, Any]) -> List[Dict[str, Any]]|None:
    """
    Обработка данных юнита для создания отчета
    
    Args:
        data: Входные данные в формате словаря
        
    Returns:
        List[Dict]: Список словарей с расчетными метриками
    """
    try:
        df = pd.DataFrame([data])
        df["Unit"] = "User"
        df["C1"] = df["customers"] / df["users"]
        df["ARPC"] = df["AVP"] * df["APC"]
        df["ARPU"] = df["ARPC"] * df["C1"]
        df["CPA"] = df["TMS"] / df["users"]
        df["CAC"] = df["TMS"] / df["customers"]
        df["CLTV"] = (df["AVP"] - df["COGS"]) * df["APC"] - df["COGS1s"]
        df["LTV"] = df["CLTV"] * df["C1"]
        df["ROI"] = (df["LTV"] - df["CPA"]) / df["CPA"] * 100
        df["UCM"] = df["LTV"] - df["CPA"]
        df["CCM"] = df["CLTV"] - df["CAC"]

        df["Profitable"] = df["UCM"] > 0

        df["Revenue"] = df["ARPU"] * df["users"]
        df["Gross_profit"] = df["CLTV"] * df["customers"]
        df["Margin"] = df["Gross_profit"] - df["TMS"]

        def calculate_required_bep(row: pd.Series):
            ucm = row.get("UCM", 0)
            if ucm > 0:
                return row.get("FC", 0) / ucm
            return None

        df["Required_units_to_BEP"] = df.apply(calculate_required_bep, axis=1)

        df["BEP"] = df["Required_units_to_BEP"] * df["UCM"]
        df["Profit"] = df["Margin"] - df["FC"]

        float_cols = df.select_dtypes(include='float').columns
        df[float_cols] = df[float_cols].round(4)

        df.replace([np.inf, -np.inf, np.nan], 0, inplace=True)
        df = df.infer_objects(copy=False)
        df = df.where(pd.notnull(df), None)[["name", "users", "customers", "AVP", "APC", "TMS", "COGS", "COGS1s", "FC", "C1", "ARPC", "ARPU", "CPA", "CAC", "CLTV", "LTV", "ROI", "UCM", "CCM", "Revenue", "Gross_profit", "Margin", "Required_units_to_BEP", "BEP", "Profit"]]
        df.columns = ["name", "users", "customers", "AVP", "APC", "TMS", "COGS", "COGS1s", "FC", "C1", "ARPC", "ARPU", "CPA", "CAC", "CLTV", "LTV", "ROI", "UCM", "CCM", "Revenue", "Gross_profit", "Margin", "Required_units_to_BEP", "BEP", "Profit"]

        return df.to_dict(orient="records")
    except Exception as e:
        logging.error(f"Ошибка в unit_calculate_economics: {e}")
        return None


def unit_count_bep(data: Dict[str, Any]) -> Tuple[Dict[str, Any], io.BytesIO]:
    """
    Рассчитывает точку безубыточности и строит график.
    
    Args:
        data: Входные данные для расчета
        
    Returns:
        Tuple[Dict, BytesIO]: Результаты расчета и изображение графика
        
    Raises:
        ValueError: Если данные некорректны или расчет невозможен
    """
    try:
        df = pd.DataFrame([data])
        proc = process_dataframe(df)
        
        if proc is None or proc.empty:
            raise ValueError("Ошибка обработки данных: результат пуст или None")
        
        record = proc.iloc[0].to_dict()
        
        required_fields = ["UCM", "FC", "Required_units_to_BEP"]
        missing_fields = [field for field in required_fields if field not in record]
        
        if missing_fields:
            raise ValueError(f"Отсутствуют обязательные поля: {missing_fields}")
        
        unit_contribution_margin = float(record["UCM"])
        fixed_costs = float(record["FC"])
        bep_units_raw = float(record["Required_units_to_BEP"])
        
        if unit_contribution_margin <= 0:
            raise ValueError("Маржинальная прибыль должна быть положительной")
        
        if fixed_costs <= 0:
            raise ValueError("Постоянные издержки должны быть положительными")
        
        if bep_units_raw < 0:
            raise ValueError("Точка безубыточности не может быть отрицательной")
        
        bep_units = int(np.ceil(bep_units_raw))
        
        max_x = max(2 * bep_units, 10)
        x_points = np.arange(0, max_x + 1)
        
        fixed_costs_line = np.full_like(x_points, -fixed_costs, dtype=float)
        
        total_profit_line = -fixed_costs + unit_contribution_margin * x_points
        
        fig, ax = plt.subplots(figsize=(10, 6))
        
        ax.plot(x_points, fixed_costs_line, 
                label="Постоянные издержки", 
                linewidth=2.5, 
                color='tab:blue',
                linestyle='--')
        
        ax.plot(x_points, total_profit_line, 
                label=f"Общая прибыль (UCM={unit_contribution_margin:.2f})", 
                linewidth=2.5, 
                color='tab:green')
        
        if 0 <= bep_units <= max_x:
            bep_profit = -fixed_costs + unit_contribution_margin * bep_units
            
            ax.axvline(x=bep_units, color='red', linestyle=':', alpha=0.7, linewidth=1.5)
            
            ax.axhline(y=bep_profit, color='red', linestyle=':', alpha=0.7, linewidth=1.5)
            
            ax.scatter(
                [bep_units], [bep_profit], 
                color='red', 
                s=100, 
                zorder=5, 
                label=f"BEP ≈ {bep_units} units",
                edgecolors='black'
                )
            annotation_text = f'BEP\n{bep_units} units\n₽{bep_profit:,.0f}'
            ax.annotate(annotation_text,
                       xy=(bep_units, bep_profit),
                       xytext=(bep_units + max_x*0.05, bep_profit + fixed_costs*0.1),
                       arrowprops=dict(arrowstyle='->', color='red', lw=1.5),
                       bbox=dict(boxstyle="round,pad=0.3", facecolor="yellow", alpha=0.8),
                       fontsize=10,
                       ha='center')
        
        if bep_units <= max_x:
            ax.fill_between(x_points[:bep_units+1], 
                           fixed_costs_line[:bep_units+1], 
                           total_profit_line[:bep_units+1],
                           where=(total_profit_line[:bep_units+1] < 0),
                           color='red', alpha=0.1, label='Зона убытков')
            
            if bep_units < len(x_points):
                ax.fill_between(x_points[bep_units:], 
                               fixed_costs_line[bep_units:], 
                               total_profit_line[bep_units:],
                               where=(total_profit_line[bep_units:] > 0),
                               color='green', alpha=0.1, label='Зона прибыли')
        
        ax.set_xlabel("Количество единиц (units)", fontsize=12)
        ax.set_ylabel("Денежный поток (₽)", fontsize=12)
        ax.set_title(f"Анализ точки безубыточности\n"
                    f"Постоянные издержки: ₽{fixed_costs:,.0f} | "
                    f"Маржинальная прибыль на единицу: ₽{unit_contribution_margin:,.2f}", 
                    fontsize=14, pad=20)
        
        ax.axhline(y=0, color='black', linewidth=1, alpha=0.5)
        ax.axvline(x=0, color='black', linewidth=1, alpha=0.5)
        
        ax.grid(True, alpha=0.3, linestyle='--')
        
        ax.legend(loc='upper left', fontsize=10, framealpha=0.9)
        
        ax.set_xlim(left=0)
        
        y_min = min(np.min(fixed_costs_line), np.min(total_profit_line))
        y_max = max(np.max(fixed_costs_line), np.max(total_profit_line))
        y_padding = abs(y_max - y_min) * 0.1
        ax.set_ylim(bottom=y_min - y_padding, top=y_max + y_padding)
        
        buf = io.BytesIO()
        fig.savefig(buf, format='png', dpi=150, bbox_inches='tight')
        buf.seek(0)
        plt.close(fig)
        
        record.update({
            "BEP_units_rounded": bep_units,
            "BEP_profit": float(bep_profit) if 'bep_profit' in locals() else 0,
            "max_units_in_chart": int(max_x)
        })
        
        return record, buf
        
    except ValueError as ve:
        logging.error(f"Ошибка валидации в unit_count_bep: {ve}")
        raise
    except Exception as e:
        logging.error(f"Непредвиденная ошибка в unit_count_bep: {e}", exc_info=True)
        raise RuntimeError(f"Ошибка при расчете точки безубыточности: {str(e)}")


@dataclass
class UnitEconomicsResult:
    """Контейнер для результатов анализа юнит-экономики"""
    basic_report: Dict[str, Any]
    bep_analysis: Dict[str, Any]
    cohort_analysis: pd.DataFrame
    files: Dict[str, io.BytesIO]
    
    def get_zip_buffer(self) -> io.BytesIO:
        """Создает ZIP архив со всеми файлами"""
        zip_buffer = io.BytesIO()
        
        with zipfile.ZipFile(zip_buffer, 'w', zipfile.ZIP_DEFLATED) as zipf:
            for filename, buffer in self.files.items():
                buffer.seek(0)
                zipf.writestr(filename, buffer.read())
        
        zip_buffer.seek(0)
        return zip_buffer


def validate_unit_economics_data(data: dict) -> Tuple[bool, str]:
    """
    Валидация входных данных для анализа юнит-экономики
    
    Args:
        data: Словарь с входными данными
        
    Returns:
        Tuple[bool, str]: (Успех валидации, Сообщение об ошибке)
    """

    required_fields = {
        'name': str,
        'users': (int, float),
        'customers': (int, float),
        'AVP': (int, float),
        'APC': (int, float),
        'TMS': (int, float),
        'COGS': (int, float),
        'COGS1s': (int, float),
        'FC': (int, float)
    }
    
    optional_fields = {
        'RR': (int, float),
        'AGR': (int, float)
    }
    
    missing_fields = []
    for field, field_type in required_fields.items():
        if field not in data:
            missing_fields.append(field)
        elif not isinstance(data[field], field_type):
            return False, f"Поле '{field}' должно быть типа {field_type}"
    
    if missing_fields:
        return False, f"Отсутствуют обязательные поля: {', '.join(missing_fields)}"
    
    validations = [
        ('users', lambda x: x > 0, "Количество пользователей должно быть > 0"),
        ('customers', lambda x: x > 0, "Количество клиентов должно быть > 0"),
        ('AVP', lambda x: x >= 0, "Средний чек должен быть ≥ 0"),
        ('APC', lambda x: x > 0, "Среднее количество покупок должно быть > 0"),
        ('TMS', lambda x: x >= 0, "Маркетинговый бюджет должен быть ≥ 0"),
        ('COGS', lambda x: x >= 0, "Себестоимость должна быть ≥ 0"),
        ('COGS1s', lambda x: x >= 0, "Дополнительные издержки должны быть ≥ 0"),
        ('FC', lambda x: x >= 0, "Постоянные издержки должны быть ≥ 0"),
    ]
    
    for field, condition, error_msg in validations:
        if field in data and not condition(data[field]):
            return False, f"{error_msg} (поле '{field}' = {data[field]})"
    

    if data.get('customers', 0) > data.get('users', 0):
        return False, f"Клиентов ({data['customers']}) не может быть больше чем пользователей ({data['users']})"
    
    return True, "Данные прошли валидацию"

def analyze_unit_economics(data: Dict[str, Any]) -> Tuple[UnitEconomicsResult, io.BytesIO]:
    """
    Полный анализ юнит-экономики с возвратом всех файлов и ZIP архива
    
    Args:
        data: Входные данные бизнес-метрик
        
    Returns:
        Tuple[UnitEconomicsResult, io.BytesIO]: Результаты и ZIP архив
    """
    try:
        success, error = validate_unit_economics_data(data)
        if not success:
            logging.error(error)
            raise ValueError(error)
        files = {}
        
        basic_result = unit_calculate_economics(data)
        if basic_result is None:
            raise ValueError("Ошибка базового расчета экономики")
        
        basic_excel_buffer = create_basic_excel_report(basic_result)
        files['basic_report.xlsx'] = basic_excel_buffer

        try:
            bep_record, bep_chart_buffer = unit_count_bep(data)
            files['bep_chart.png'] = bep_chart_buffer
            
            bep_df = pd.DataFrame([bep_record])
            bep_excel_buffer = create_excel_buffer(bep_df, "BEP_Анализ")
            files['bep_analysis.xlsx'] = bep_excel_buffer
        except Exception as bep_error:
            logging.warning(f"Ошибка при расчете BEP: {bep_error}")
            bep_record = {"error": str(bep_error)}
        
        try:
            cohort_buffer = unit_count_cohort_raw(data)
            files['cohort_analysis.xlsx'] = cohort_buffer
            
            cohort_df = get_cohort_dataframe(data)
            
            cohort_images = create_cohort_charts(cohort_df)
            for i, (name, img_buffer) in enumerate(cohort_images.items(), 1):
                files[f'cohort_{name}.png'] = img_buffer
                
        except Exception as cohort_error:
            logging.warning(f"Ошибка при когортном анализе: {cohort_error}")
            cohort_df = pd.DataFrame()
        
        summary_buffer = create_summary_report(basic_result, bep_record, cohort_df)
        files['summary_report.txt'] = summary_buffer
        
        json_buffer = create_json_export(basic_result, bep_record, cohort_df)
        files['full_data.json'] = json_buffer
        
        consolidated_excel_buffer = create_consolidated_excel(
            basic_result, bep_record, cohort_df
        )
        files['consolidated_report.xlsx'] = consolidated_excel_buffer
        
        result = UnitEconomicsResult(
            basic_report=basic_result[0] if basic_result else {},
            bep_analysis=bep_record,
            cohort_analysis=cohort_df,
            files=files
        )
        
        zip_buffer = result.get_zip_buffer()
        
        return result, zip_buffer
        
    except Exception as e:
        logging.error(f"Ошибка в полном анализе юнит-экономики: {e}", exc_info=True)
        raise


def create_basic_excel_report(basic_result: List[Dict]) -> io.BytesIO:
    """Создает Excel файл с базовым отчетом"""
    df = pd.DataFrame(basic_result)
    return create_excel_buffer(df, "Основные_метрики")


def create_excel_buffer(df: pd.DataFrame, sheet_name: str) -> io.BytesIO:
    """Создает буфер Excel файла с форматированием"""
    buffer = io.BytesIO()
    
    with pd.ExcelWriter(buffer, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name=sheet_name)
        
        worksheet = writer.sheets[sheet_name]
        header_font = Font(bold=True)
        
        for col_num, column_title in enumerate(df.columns, 1):
            cell = worksheet.cell(row=1, column=col_num)
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
            
        for column_cells in worksheet.columns:
            length = max(len(str(cell.value)) if cell.value else 0 for cell in column_cells)
            col_letter = get_column_letter(column_cells[0].column)
            worksheet.column_dimensions[col_letter].width = min(length + 2, 50)
    
    buffer.seek(0)
    return buffer


def unit_count_cohort_raw(data: Dict[str, Any]) -> io.BytesIO:
    """Упрощенная версия когортного анализа, возвращающая только Excel"""
    try:
        result = pd.DataFrame([data])
        
        required_fields = ['RR', 'AGR', 'FC']
        for field in required_fields:
            if field not in result.columns:
                result[field] = 0.1
        
        conv = (result["customers"] / result["users"]).iloc[0] if result["users"].iloc[0] > 0 else 0.1

        result["cohort"] = 1
        result["new users"] = result["users"]
        result.loc[0, "new users"] = result.loc[0, "users"]
        result["user retention"] = 0
        result["user churn"] = 0
        result["total users"] = result.loc[0, "users"]
        result["Accumulative profit"] = 0
        result["Ballance"] = 0 - result["FC"].iloc[0]

        periods = 24
        expanded_rows = []
        
        for i in range(periods):
            row = result.iloc[0].copy()
            row["cohort"] = i + 1
            
            if i == 0:
                row["new users"] = row["users"]
                row["user retention"] = 0
                row["user churn"] = 0
                row["total users"] = row["users"]
                row["Ballance"] = -row["FC"]
            else:
                prev_row = expanded_rows[i-1]
                rr = row.get("RR", 0.7)
                agr = row.get("AGR", 0.05)
                
                row["new users"] = prev_row["total users"] * agr
                row["user retention"] = prev_row["total users"] * rr
                row["user churn"] = prev_row["total users"] * (1 - rr)
                row["total users"] = row["user retention"] + row["new users"]
                row["customers"] = row["total users"] * conv
                row["C1"] = conv
                row["ARPC"] = row["AVP"] * row["APC"]
                row["ARPU"] = row["ARPC"] * row["C1"]
                row["CPA"] = row["TMS"] / row["total users"] if row["total users"] > 0 else 0
                row["CAC"] = row["TMS"] / row["customers"] if row["customers"] > 0 else 0
                row["CLTV"] = (row["AVP"] - row["COGS"]) * row["APC"] - row["COGS1s"]
                row["LTV"] = row["CLTV"] * row["C1"]
                row["ROI"] = (row["LTV"] - row["CPA"]) / row["CPA"] * 100 if row["CPA"] > 0 else 0
                row["UCM"] = row["LTV"] - row["CPA"]
                row["CCM"] = row["CLTV"] - row["CAC"]
                row["Revenue"] = row["ARPU"] * row["total users"]
                row["Gross_profit"] = row["CLTV"] * row["customers"]
                row["Margin"] = row["Gross_profit"] - row["TMS"]
                
                row["Profit"] = row["Margin"] - row["FC"]
                row["Accumulative profit"] = prev_row["Accumulative profit"] + row["Profit"]
                row["Ballance"] = prev_row["Ballance"] + row["Profit"] - row["FC"]
            
            expanded_rows.append(row)
        
        result = pd.DataFrame(expanded_rows)
        
        float_cols = result.select_dtypes(include='float').columns
        result[float_cols] = result[float_cols].round(2)
        result.replace([np.inf, -np.inf], np.nan, inplace=True)
        result = result.fillna(0)
        
        return create_excel_buffer(result, "Когортный_анализ")
        
    except Exception as e:
        logging.error(f"Ошибка в когортном анализе: {e}")
        return io.BytesIO()


def get_cohort_dataframe(data: Dict[str, Any]) -> pd.DataFrame:
    """Получает DataFrame для когортного анализа"""
    buffer = unit_count_cohort_raw(data)
    buffer.seek(0)
    
    try:
        df = pd.read_excel(buffer, sheet_name=0)
        return df
    except:
        return pd.DataFrame()


def create_cohort_charts(cohort_df: pd.DataFrame) -> Dict[str, io.BytesIO]:
    """Создает графики для когортного анализа"""
    charts = {}
    
    if cohort_df.empty:
        return charts
    
    try:
        if 'Profit' in cohort_df.columns:
            fig, ax = plt.subplots(figsize=(10, 6))
            ax.plot(cohort_df['cohort'], cohort_df['Profit'], 
                   marker='o', linewidth=2, color='green')
            ax.set_xlabel('Период (месяц)')
            ax.set_ylabel('Прибыль')
            ax.set_title('Динамика прибыли по периодам')
            ax.grid(True, alpha=0.3)
            ax.axhline(y=0, color='red', linestyle='--', alpha=0.5)
            
            buf = io.BytesIO()
            fig.savefig(buf, format='png', dpi=150, bbox_inches='tight')
            buf.seek(0)
            charts['profit_dynamics.png'] = buf
            plt.close(fig)
        
        # 2. График аудитории
        if 'total users' in cohort_df.columns:
            fig, ax = plt.subplots(figsize=(10, 6))
            ax.plot(cohort_df['cohort'], cohort_df['total users'], 
                   marker='s', linewidth=2, color='blue')
            ax.set_xlabel('Период (месяц)')
            ax.set_ylabel('Общая аудитория')
            ax.set_title('Рост аудитории')
            ax.grid(True, alpha=0.3)
            
            buf = io.BytesIO()
            fig.savefig(buf, format='png', dpi=150, bbox_inches='tight')
            buf.seek(0)
            charts['audience_growth.png'] = buf
            plt.close(fig)
        
        if 'Accumulative profit' in cohort_df.columns:
            fig, ax = plt.subplots(figsize=(10, 6))
            ax.plot(cohort_df['cohort'], cohort_df['Accumulative profit'], 
                   marker='^', linewidth=2, color='purple')
            ax.set_xlabel('Период (месяц)')
            ax.set_ylabel('Накопленная прибыль')
            ax.set_title('Накопленная прибыль')
            ax.grid(True, alpha=0.3)
            ax.axhline(y=0, color='red', linestyle='--', alpha=0.5)
            
            buf = io.BytesIO()
            fig.savefig(buf, format='png', dpi=150, bbox_inches='tight')
            buf.seek(0)
            charts['accumulated_profit.png'] = buf
            plt.close(fig)
            
    except Exception as e:
        logging.error(f"Ошибка при создании графиков: {e}")
    
    return charts


def create_summary_report(basic_result: List[Dict], bep_analysis: Dict, 
                         cohort_df: pd.DataFrame) -> io.BytesIO:
    """Создает текстовый сводный отчет"""
    
    buffer = io.BytesIO()
    
    report_lines = [
        "=" * 60,
        "ОТЧЕТ ПО ЮНИТ-ЭКОНОМИКЕ",
        f"Дата генерации: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}",
        "=" * 60,
        "",
        "1. ОСНОВНЫЕ МЕТРИКИ:",
        "-" * 40,
    ]
    
    if basic_result:
        br = basic_result[0]
        report_lines.extend([
            f"Название: {br.get('name', 'N/A')}",
            f"Пользователи: {br.get('users', 0):,.0f}",
            f"Клиенты: {br.get('customers', 0):,.0f}",
            f"Конверсия (C1): {br.get('C1', 0):.2%}",
            f"ARPU: ₽{br.get('ARPU', 0):,.2f}",
            f"CAC: ₽{br.get('CAC', 0):,.2f}",
            f"LTV: ₽{br.get('LTV', 0):,.2f}",
            f"ROI: {br.get('ROI', 0):.1f}%",
            f"UCM: ₽{br.get('UCM', 0):,.2f}",
            f"Прибыль: ₽{br.get('Profit', 0):,.2f}",
        ])
    
    report_lines.extend([
        "",
        "2. ТОЧКА БЕЗУБЫТОЧНОСТИ (BEP):",
        "-" * 40,
    ])
    
    if bep_analysis and 'error' not in bep_analysis:
        report_lines.extend([
            f"Точка безубыточности: {bep_analysis.get('BEP_units_rounded', 0):,.0f} юнитов",
            f"Постоянные издержки (FC): ₽{bep_analysis.get('FC', 0):,.2f}",
            f"Маржинальная прибыль на юнит (UCM): ₽{bep_analysis.get('UCM', 0):,.2f}",
        ])
    
    report_lines.extend([
        "",
        "3. КОГОРТНЫЙ АНАЛИЗ (24 периода):",
        "-" * 40,
    ])
    
    if not cohort_df.empty:
        report_lines.extend([
            f"Итого периодов: {len(cohort_df)}",
            f"Общая аудитория (конечная): {cohort_df['total users'].iloc[-1]:,.0f}",
            f"Накопленная прибыль: ₽{cohort_df['Accumulative profit'].iloc[-1]:,.2f}",
            f"Финальный баланс: ₽{cohort_df['Ballance'].iloc[-1]:,.2f}",
        ])
    
    report_lines.extend([
        "",
        "=" * 60,
        "ВЫВОДЫ:",
        "-" * 40,
    ])
    
    if basic_result:
        br = basic_result[0]
        if br.get('ROI', 0) > 100:
            report_lines.append("✓ Высокая рентабельность (ROI > 100%)")
        elif br.get('ROI', 0) > 0:
            report_lines.append("✓ Положительная рентабельность")
        else:
            report_lines.append("✗ Отрицательная рентабельность")
        
        if br.get('UCM', 0) > 0:
            report_lines.append("✓ Положительная маржинальная прибыль на юнит")
        else:
            report_lines.append("✗ Отрицательная маржинальная прибыль на юнит")
    
    report_text = "\n".join(report_lines)
    buffer.write(report_text.encode('utf-8'))
    buffer.seek(0)
    
    return buffer


def create_json_export(
    basic_result: List[Dict], 
    bep_analysis: Dict, 
    cohort_df: pd.DataFrame) -> io.BytesIO:
    """Создает JSON файл с полными данными"""
    import json
    
    export_data = {
        "timestamp": datetime.now().isoformat(),
        "basic_metrics": basic_result[0] if basic_result else {},
        "bep_analysis": bep_analysis,
        "cohort_summary": {
            "total_periods": len(cohort_df),
            "final_audience": cohort_df['total users'].iloc[-1] if not cohort_df.empty else 0,
            "total_profit": cohort_df['Accumulative profit'].iloc[-1] if not cohort_df.empty else 0,
            "final_balance": cohort_df['Ballance'].iloc[-1] if not cohort_df.empty else 0
        } if not cohort_df.empty else {}
    }
    
    buffer = io.BytesIO()
    
    json_str = json.dumps(export_data, ensure_ascii=False, indent=2, default=str)
    buffer.write(json_str.encode('utf-8'))
    buffer.seek(0)
    
    return buffer


def create_consolidated_excel(basic_result: List[Dict], bep_analysis: Dict,
                             cohort_df: pd.DataFrame) -> io.BytesIO:
    """Создает единую Excel книгу со всеми листами"""
    buffer = io.BytesIO()
    
    with pd.ExcelWriter(buffer, engine='openpyxl') as writer:
        if basic_result:
            basic_df = pd.DataFrame(basic_result)
            basic_df.to_excel(writer, sheet_name='Основные_метрики', index=False)
        
        if bep_analysis and 'error' not in bep_analysis:
            bep_df = pd.DataFrame([bep_analysis])
            bep_df.to_excel(writer, sheet_name='BEP_Анализ', index=False)
        
        if not cohort_df.empty:
            cohort_df.to_excel(writer, sheet_name='Когортный_анализ', index=False)
        
        summary_data = {
            'Метрика': ['Дата отчета', 'Статус', 'Рекомендации'],
            'Значение': [
                datetime.now().strftime('%Y-%m-%d %H:%M'),
                'Успешно сгенерировано' if basic_result else 'Ошибка',
                'Проанализируйте точки роста и оптимизации' if basic_result else 'Требуется проверка данных'
            ]
        }
        summary_df = pd.DataFrame(summary_data)
        summary_df.to_excel(writer, sheet_name='Итоги', index=False)
    
    buffer.seek(0)
    return buffer


def prepare_files_for_bot(economics_result: UnitEconomicsResult) -> List[Tuple[str, io.BytesIO]]:
    """
    Подготавливает файлы для отправки ботом
    
    Returns:
        List[Tuple[filename, buffer]]: Список файлов для отправки
    """
    files = []
    
    key_files = ['basic_report.xlsx', 'bep_chart.png', 'summary_report.txt']
    
    for filename in key_files:
        if filename in economics_result.files:
            buffer = economics_result.files[filename]
            buffer.seek(0)
            files.append((filename, buffer))
    
    zip_buffer = economics_result.get_zip_buffer()
    files.append(('unit_economics_full_package.zip', zip_buffer))
    
    return files



if __name__ == "__main__":
    sample_data = {
        "name": "Test Product",
        "users": 1000,
        "customers": 100,
        "AVP": 50.0,
        "APC": 2.0,
        "TMS": 5000.0,
        "COGS": 15.0, 
        "COGS1s": 5.0, 
        "FC": 10000.0,
        "RR": 0.8, 
        "AGR": 0.05
    }
    
    try:
        result, zip_buffer = analyze_unit_economics(sample_data)
        
        print("✅ Анализ успешно завершен!")
        print(f"📊 Основные метрики: {len(result.basic_report)} показателей")
        print(f"📈 BEP анализ: {'успешно' if 'error' not in result.bep_analysis else 'с ошибкой'}")
        print(f"📋 Когортный анализ: {len(result.cohort_analysis)} периодов")
        print(f"📁 Файлов создано: {len(result.files)}")
        print(f"🗜️  Размер ZIP архива: {len(zip_buffer.getvalue())} байт")
        
        with open("unit_economics_report.zip", "wb") as f:
            f.write(zip_buffer.getvalue())
        print("💾 ZIP архив сохранен как 'unit_economics_report.zip'")
        
    except Exception as e:
        print(f"❌ Ошибка при анализе: {e}")